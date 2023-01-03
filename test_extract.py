import os
import pathlib
import zipfile
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import pytest

files_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')
resources_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
archive_path = os.path.join(resources_path, 'test.zip')


@pytest.fixture(scope='function', autouse=False)
def archived_files():
    with zipfile.ZipFile(archive_path, mode='w') as zip_file:
        for file in pathlib.Path(files_path).iterdir():
            zip_file.write(file, arcname=file.name)
    assert len(os.listdir(resources_path)) == 1

    yield

    os.remove(archive_path)
    assert len(os.listdir(resources_path)) == 0


# read csv file
def test_read_csv(archived_files):
    with zipfile.ZipFile(archive_path) as c_f:
        csv_archived = c_f.extract('test.csv')
        with open(csv_archived) as csv_file:
            csv_rows = csv.reader(csv_file)
            new_list = []
            for row in csv_rows:
                new_list.append(row)
            assert len(new_list) == 2
            assert new_list == [['Имя клиента;Телефон клиента'],
                                ['test;test1']]
        os.remove('test.csv')


# read pdf file
def test_read_pdf(archived_files):
    with zipfile.ZipFile(archive_path) as p_f:
        pdf_file = 'test.pdf'
        pdf_archived = p_f.extract(pdf_file)
        reader = PdfReader(pdf_archived)
        assert len(reader.pages) == 1
        page = reader.pages[0]
        text = page.extract_text()
        assert 'Sprinkled with Selenium usage tips' in text
        os.remove(pdf_file)


# read xlsx file
def test_read_xlsx(archived_files):
    with zipfile.ZipFile(archive_path) as x_f:
        xlsx_file = 'test.xlsx'
        xlsx_archived = x_f.extract(xlsx_file)
        workbook = load_workbook(xlsx_archived)
        sheet = workbook.active
        assert sheet.cell(row=1, column=1).value == 'Имя клиента'
        os.remove(xlsx_file)
